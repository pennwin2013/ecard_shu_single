#!/usr/bin/python
# -*- coding: gbk -*-

import sys,os,glob
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.style import Color
from openpyxl.writer.styles import StyleWriter

class doc_reader:
    def __init__(self,docfile):
        self.doc_file = docfile
        self.doc_name = os.path.basename(self.doc_file)
        self.sheet_name = u'接口表'
        self.result_list = []
        
    def cell_value(self,cell):
        if cell is None:
            return None
        #print type(cell.value)
        if isinstance(cell.value,unicode):
            return cell.value.encode('gbk')
        elif isinstance(cell.value,int):
            return '%d' % cell.value
        elif isinstance(cell.value,float):
            return '%f' % cell.value
        else:
            return cell.value
        
    def cell_name(self,cell,no):
        cn = '%s%d' % (cell,no)
        return cn
        
    def read(self):
        print "读取[%s]..." % self.doc_name
        self.result_list = []
        wb = load_workbook(self.doc_file)
        if self.sheet_name not in wb.get_sheet_names():
            print "没有定义Sheet[%s]" % self.sheet_name
            return False
        ws = wb.get_sheet_by_name(self.sheet_name)
        
        rowno = 1
        empty_rows = 0
        interface_count = 0
        pre_cell_value = ''
        while True:
            cell = ws.cell(self.cell_name('A',rowno))
            if cell is None:
                break
            cell_value = self.cell_value(cell)
            if cell_value == '功能号':
                funcno = self.cell_value(ws.cell(self.cell_name('B',rowno)))
                print '读取接口：%s : %s' % (funcno,pre_cell_value)
                self.result_list.append((funcno,pre_cell_value,self.doc_name,rowno-1))
                empty_rows = 0
                interface_count += 1
            elif cell_value == 'END_DEFINE':
                print '读取[%s]完成，接口[%d]个\n' % (self.doc_name,interface_count)
                return True
            elif cell_value == '' or cell_value == None:
                empty_rows += 1
                if empty_rows > 100:
                    print '读取[%s]完成，接口[%d]个\n' % (self.doc_name,interface_count)
                    break
            rowno += 1
            pre_cell_value = cell_value

        return True
    
    def result(self):
        return self.result_list
        

class create_list:
    def __init__(self):
        self.dir = '.'
        self.exec_name = ''
        self.exec_path = ''
        self.output_file = '00.功能目录.xlsx'
        self.output = None
        self.out_sheet = None
        self.sheet_row = 2
        
    def create_output(self):
        full_output = os.path.join(self.dir,self.output_file)
        if os.path.exists(full_output):
            ans = raw_input('文件[%s]已存在，是否覆盖[(y)es/(n)o]?' % full_output)
            if ans not in ('y','Y'):
                return False
            os.remove(full_output)
                
        self.output = Workbook()
        self.out_sheet = self.output.create_sheet(0)
        self.out_sheet.title = 'Sheet1'
        
        self.out_sheet.cell('A1').value = u'功能号'
        self.out_sheet.cell('A1').style.fill.fill_type = 'fill'
        self.out_sheet.cell('A1').style.fill.start_color.index = Color.DARKYELLOW
        self.out_sheet.cell('B1').value = u'描述'
        #self.out_sheet.cell('B1').style.fill.fill_type = 'darkGray'
        self.out_sheet.cell('C1').value = u'文档名称'
        #self.out_sheet.cell('C1').style.fill.fill_type = 'darkGray'
        self.sheet_row = 2
        return True
        
    def save_output(self):
        if self.output is None:
            return
        full_output = os.path.join(self.dir,self.output_file)
        self.output.save(full_output)
    
    def value_2_cell(self,value):
        if isinstance(value,str):
            return unicode(value,'gbk')
        else:
            return value
            
    def append_item(self,item):
        cellname = 'A%d' % self.sheet_row
        cell = self.out_sheet.cell(cellname)
        cell.value = self.value_2_cell(item[0])
        
        cellname = 'B%d' % self.sheet_row
        cell = self.out_sheet.cell(cellname)
        cell.value = self.value_2_cell(item[1])
        
        cellname = 'C%d' % self.sheet_row
        cell = self.out_sheet.cell(cellname)
        cell.value = self.value_2_cell(item[2])
        full_name = os.path.join(self.dir,item[2])
        
        #hyperlink = 'file:///' + full_name + '/接口表!A%d' % item[3]
        #print self.value_2_cell(hyperlink)
        #cell.hyperlink = hyperlink
        
        self.sheet_row += 1
        
        return True
        
    def run(self):
        if len(sys.argv) > 1:
            self.dir = sys.argv[1]
            
        self.exec_name = os.path.basename(sys.argv[0])
        self.exec_path = os.path.dirname(os.path.realpath(sys.argv[0]))
        
        if not os.path.exists(self.dir):
            print "读取目录[%s]失败" % self.dir
            sys.exit(1)
            
        self.dir = os.path.realpath(self.dir)
        if not self.create_output():
            return
        
        doc_files = glob.glob(os.path.join(self.dir,'*.xlsx'))
        for doc in doc_files:
            doc = os.path.basename(doc)
            if doc == '00.前后台接口-模板.xlsx': continue
            if doc.startswith('~'): continue
            reader = doc_reader(os.path.join(self.dir,doc))
            if not reader.read():
                ans = raw_input('是否继续[(e)xit/(c)ontinue]?')
                if ans in ('e','E'):
                    break
            # continue
            for item in reader.result():
                if not self.append_item(item):
                    print '保存文件目录失败'
                    sys.exit(1)
                
        self.save_output()
        
        
if __name__ == "__main__":
    c = create_list()
    c.run()
    